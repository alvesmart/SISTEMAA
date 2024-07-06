import os
import tkinter as tk
from tkinter import ttk, messagebox
import pandas as pd
from datetime import datetime
from tkinter.filedialog import asksaveasfilename
from openpyxl import load_workbook
from openpyxl.styles.alignment import Alignment
import ttkbootstrap as tb
from ttkbootstrap.constants import *

# Arquivo para salvar os dados
ARQUIVO_DADOS = "registros.csv"

# Inicializar DataFrame para armazenar registros
columns = ["Nome Completo", "Identidade", "Modelo do Carro", "Placa", "Endereço", "Número",
           "Data de Entrada", "Hora de Entrada", "Data de Saída", "Hora de Saída", "Tipo"]
df = pd.DataFrame(columns=columns)

# Verificar se o arquivo existe e carregar dados se sim
if os.path.isfile(ARQUIVO_DADOS):
    df = pd.read_csv(ARQUIVO_DADOS, index_col=0)

# Lista de opções para o campo Tipo
tipos_opcoes = ["Visitante", "Morador", "Funcionário"]

# Função para salvar os registros ao fechar o programa
def salvar_registros():
    df.to_csv(ARQUIVO_DADOS)

# Função para verificar campos vazios
def verificar_campos_vazios():
    campos = [
        nome_entry.get(),
        identidade_entry.get(),
        modelo_carro_entry.get(),
        placa_entry.get(),
        endereco_entry.get(),
        numero_entry.get(),
        tipo_combobox.get()
    ]
    return all(campos)

# Função para verificar se o nome já existe nos registros
def verificar_nome_existente(nome):
    return nome in df["Nome Completo"].values

# Função para registrar entrada
def registrar_entrada(nome, identidade, modelo_carro, placa, endereco, numero, tipo):
    global df
    if verificar_campos_vazios():
        if verificar_nome_existente(nome):
            messagebox.showwarning("Aviso", f"Já existe um registro para o nome '{nome}'.")
        else:
            now = datetime.now()
            data_entrada = now.strftime("%d-%m-%Y")
            hora_entrada = now.strftime("%H:%M:%S")
            novo_registro = pd.DataFrame([[nome, identidade, modelo_carro, placa, endereco, numero, data_entrada,
                                           hora_entrada, "", "", tipo]], columns=columns)
            df = pd.concat([df, novo_registro], ignore_index=True)
            messagebox.showinfo("Sucesso", "Entrada registrada com sucesso!")
            limpar_campos_entrada()
            atualizar_lista(tree)
            centralizar_janela()
    else:
        messagebox.showwarning("Aviso", "Todos os campos devem ser preenchidos.")

# Função para registrar saída
def registrar_saida():
    global df
    selected_items = tree.selection()
    if selected_items:
        for item in selected_items:
            iid = int(item)
            if df.at[iid, "Data de Saída"] == "":
                now = datetime.now()
                df.at[iid, "Data de Saída"] = now.strftime("%d-%m-%Y")
                df.at[iid, "Hora de Saída"] = now.strftime("%H:%M:%S")
                messagebox.showinfo("Sucesso", "Saída registrada com sucesso!")
                atualizar_lista(tree)
            else:
                messagebox.showwarning("Aviso", f"Saída já registrada para o registro de ID {iid}.")
    else:
        messagebox.showwarning("Aviso", "Nenhum registro selecionado para registrar saída.")

# Função para exportar para Excel
def exportar_para_excel():
    nome_arquivo = asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    if nome_arquivo:
        df.to_excel(nome_arquivo, index=False)
        
        # Abrir o arquivo Excel para centralizar as células
        wb = load_workbook(nome_arquivo)
        sheet = wb.active
        
        # Centralizar todas as células
        for row in sheet.iter_rows(min_row=2, min_col=1, max_row=sheet.max_row, max_col=sheet.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        wb.save(nome_arquivo)
        wb.close()
        
        messagebox.showinfo("Sucesso", f"Registros exportados para {nome_arquivo} com sucesso!")

# Função para atualizar a lista na Treeview
def atualizar_lista(tree):
    tree.delete(*tree.get_children())
    df_sorted = df.sort_values(by=["Data de Entrada", "Hora de Entrada"], ascending=False)
    for index, row in df_sorted.iterrows():
        tree.insert("", "end", iid=index, values=row.tolist())

# Função para excluir registro selecionado
def excluir_registro_selecionado():
    global df
    selected_items = tree.selection()
    if selected_items:
        confirmar = messagebox.askyesno("Confirmação", "Tem certeza de que deseja excluir os registros selecionados?")
        if confirmar:
            for item in selected_items:
                iid = int(item)
                df.drop(index=iid, inplace=True)
            df.reset_index(drop=True, inplace=True)
            atualizar_lista(tree)
            messagebox.showinfo("Sucesso", "Registros selecionados excluídos com sucesso.")
    else:
        messagebox.showwarning("Aviso", "Nenhum registro selecionado para exclusão.")

# Função para pesquisar registros
def pesquisar_registros(termo):
    tree.delete(*tree.get_children())
    df_filtered = df[df["Nome Completo"].str.contains(termo, case=False, na=False)]
    for index, row in df_filtered.iterrows():
        tree.insert("", "end", iid=index, values=row.tolist())

# Função para limpar campos de entrada
def limpar_campos_entrada():
    nome_entry.delete(0, tk.END)
    identidade_entry.delete(0, tk.END)
    modelo_carro_entry.delete(0, tk.END)
    placa_entry.delete(0, tk.END)
    endereco_entry.delete(0, tk.END)
    numero_entry.delete(0, tk.END)
    tipo_combobox.set("")

# Função para centralizar a janela
def centralizar_janela():
    root.update_idletasks()
    width = root.winfo_width()
    height = root.winfo_height()
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x = (screen_width // 2) - (width // 2)
    y = (screen_height // 2) - (height // 2)
    root.geometry(f"{width}x{height}+{x}+{y}")

# Função para fechar o programa
def fechar_programa():
    salvar_registros()
    root.destroy()

# Configurar GUI com ttkbootstrap
root = tb.Window(themename="journal")
root.title("Sistema de Controle")
root.geometry("1000x650")

# Barra de menu
menubar = tk.Menu(root)
root.config(menu=menubar)

arquivo_menu = tk.Menu(menubar, tearoff=0)
arquivo_menu.add_command(label="Exportar para Excel", command=exportar_para_excel)
arquivo_menu.add_separator()
arquivo_menu.add_command(label="Sair", command=fechar_programa)
menubar.add_cascade(label="Arquivo", menu=arquivo_menu)

# Criando abas
tabControl = ttk.Notebook(root)
tab_registro = ttk.Frame(tabControl)
tab_lista = ttk.Frame(tabControl)
tabControl.add(tab_registro, text="Registro")
tabControl.add(tab_lista, text="Lista")
tabControl.pack(expand=1, fill="both", padx=10, pady=10)

# Elementos na aba Registro
ttk.Label(tab_registro, text="Nome Completo:").grid(column=0, row=0, padx=10, pady=5)
ttk.Label(tab_registro, text="Identidade:").grid(column=0, row=1, padx=10, pady=5)
ttk.Label(tab_registro, text="Modelo do Carro:").grid(column=0, row=2, padx=10, pady=5)
ttk.Label(tab_registro, text="Placa:").grid(column=0, row=3, padx=10, pady=5)
ttk.Label(tab_registro, text="Endereço:").grid(column=0, row=4, padx=10, pady=5)
ttk.Label(tab_registro, text="Número:").grid(column=0, row=5, padx=10, pady=5)
ttk.Label(tab_registro, text="Tipo:").grid(column=0, row=6, padx=10, pady=5)

nome_entry = ttk.Entry(tab_registro)
identidade_entry = ttk.Entry(tab_registro)
modelo_carro_entry = ttk.Entry(tab_registro)
placa_entry = ttk.Entry(tab_registro)
endereco_entry = ttk.Entry(tab_registro)
numero_entry = ttk.Entry(tab_registro)

tipo_combobox = ttk.Combobox(tab_registro, values=tipos_opcoes, state="readonly")
tipo_combobox.grid(column=1, row=6, padx=10, pady=5)

nome_entry.grid(column=1, row=0, padx=10, pady=5)
identidade_entry.grid(column=1, row=1, padx=10, pady=5)
modelo_carro_entry.grid(column=1, row=2, padx=10, pady=5)
placa_entry.grid(column=1, row=3, padx=10, pady=5)
endereco_entry.grid(column=1, row=4, padx=10, pady=5)
numero_entry.grid(column=1, row=5, padx=10, pady=5)

def registrar_entrada_button():
    registrar_entrada(
        nome_entry.get(),
        identidade_entry.get(),
        modelo_carro_entry.get(),
        placa_entry.get(),
        endereco_entry.get(),
        numero_entry.get(),
        tipo_combobox.get()
    )

ttk.Button(tab_registro, text="Registrar Entrada", style="primary.TButton", command=registrar_entrada_button).grid(column=0, row=7, columnspan=2, padx=10, pady=10)

# Elementos na aba Lista
tree = ttk.Treeview(tab_lista, columns=columns, show="headings", style="primary.Treeview")
for col in columns:
    tree.heading(col, text=col)
    tree.column(col, width=120, anchor="center")
tree.pack(expand=True, fill="both", padx=10, pady=10)

# Botões adicionais na aba Lista
frame_botoes = ttk.Frame(tab_lista)
frame_botoes.pack(pady=10)

ttk.Button(frame_botoes, text="Registrar Saída", style="success.TButton", command=registrar_saida).grid(row=0, column=0, padx=10)
ttk.Button(frame_botoes, text="Excluir Selecionado", style="danger.TButton", command=excluir_registro_selecionado).grid(row=0, column=1, padx=10)
ttk.Button(frame_botoes, text="Exportar para Excel", style="info.TButton", command=exportar_para_excel).grid(row=0, column=2, padx=10)

# Campo de pesquisa
pesquisa_frame = ttk.Frame(tab_lista)
pesquisa_frame.pack(pady=10)

ttk.Label(pesquisa_frame, text="Pesquisar:").grid(row=0, column=0, padx=10)
pesquisa_entry = ttk.Entry(pesquisa_frame)
pesquisa_entry.grid(row=0, column=1, padx=10)

def pesquisar_button_click():
    termo = pesquisa_entry.get().strip()
    pesquisar_registros(termo)

pesquisar_button = ttk.Button(pesquisa_frame, text="Pesquisar", style="primary.TButton", command=pesquisar_button_click)
pesquisar_button.grid(row=0, column=2, padx=10)

# Centralizar a janela
root.after(0, centralizar_janela)

# Lidar com o fechamento da janela
root.protocol("WM_DELETE_WINDOW", fechar_programa)

root.mainloop()